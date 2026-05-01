#!/usr/bin/env python3
"""
Generate a Bromyard depot haulier packing sheet XLSX.

Confirmed v2 layout — do not alter structure without Gareth's approval.

Usage:
    python generate_packing_sheet.py \
        --manifest manifest.csv \
        --bale-log bale-log.csv \
        --load-number 5 \
        --haulier "Nigel" \
        --prev-total 4821.0 \
        --date 01/05/2026 \
        --output Packing_Sheet_Load5.xlsx
"""

import argparse
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Colours ──────────────────────────────────────────────────────────────────

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREY   = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

BORDER = _thin_border()

# ── CSV loaders ───────────────────────────────────────────────────────────────

def load_bale_log(path):
    """Return dict keyed by bale number string → row dict."""
    bales = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = str(row["Bale No."]).strip()
            bales[key] = row
    return bales


def load_manifest(path):
    """Return list of row dicts from depot suite CSV."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# ── Cross-reference ───────────────────────────────────────────────────────────

def cross_reference(manifest_rows, bale_log):
    """
    Match manifest bales against the bale log.

    Returns (valid_rows, errors).
    Stops if any bale is missing from the log — caller should abort on hard errors.
    """
    valid  = []
    errors = []

    for row in manifest_rows:
        bale_no       = str(row["Bale Number"]).strip()
        manifest_type = str(row.get("Type", "")).strip()

        if bale_no not in bale_log:
            errors.append((bale_no, "HARD", "no matching entry in bale log"))
            continue

        log = bale_log[bale_no]

        if log.get("Transported", "FALSE").strip().upper() == "TRUE":
            errors.append((bale_no, "SKIP", "already marked as Transported"))
            continue

        log_type = log.get("Type", "").strip()
        if log_type and manifest_type and log_type != manifest_type:
            errors.append((
                bale_no, "HARD",
                f"type mismatch — manifest '{manifest_type}' vs log '{log_type}'"
            ))
            continue

        try:
            gross = float(log["Gross Weight (kg)"])
            bags  = int(log["No. of Sheets"])
        except (ValueError, KeyError) as exc:
            errors.append((bale_no, "HARD", f"weight data missing or invalid: {exc}"))
            continue

        valid.append({
            "bale_no": bale_no,
            "type":    manifest_type or log_type,
            "gross":   gross,
            "bags":    bags,
            "net":     gross - bags,
        })

    return valid, errors


# ── Comments column logic ─────────────────────────────────────────────────────

_CLIP_SIZE  = {"2BC": 2, "3BC": 3, "4BC": 4}
_CLIP_LABEL = {"2BC": "2 BALE CLIP", "3BC": "3 BALE CLIP", "4BC": "4 BALE CLIP"}

def build_comments(rows):
    """
    Add 'comments' field to each row.

    Singles are blank. First bale of a clip gets full label ("2 BALE CLIP" etc.),
    subsequent bales in the same clip get a ditto mark.
    """
    result = []
    i = 0
    while i < len(rows):
        btype = rows[i]["type"]
        if btype == "Single":
            result.append({**rows[i], "comments": ""})
            i += 1
        else:
            size  = _CLIP_SIZE.get(btype, 1)
            label = _CLIP_LABEL.get(btype, btype)
            for j in range(size):
                if i + j < len(rows):
                    result.append({**rows[i + j], "comments": label if j == 0 else '"'})
            i += size
    return result


# ── Cell helpers ──────────────────────────────────────────────────────────────

def _set(ws, ref, value=None, bold=False, italic=False, size=None,
         fill=None, border=False, halign=None, valign=None, wrap=False):
    cell = ws[ref]
    if value is not None:
        cell.value = value
    font_kw = dict(bold=bold, italic=italic)
    if size:
        font_kw["size"] = size
    cell.font = Font(**font_kw)
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    align_kw = {}
    if halign:
        align_kw["horizontal"] = halign
    if valign:
        align_kw["vertical"] = valign
    if wrap:
        align_kw["wrap_text"] = True
    if align_kw:
        cell.alignment = Alignment(**align_kw)
    return cell


# ── XLSX generation ───────────────────────────────────────────────────────────

def generate_xlsx(rows, load_number, haulier, prev_total, date_str, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Load {load_number}"

    # Column widths: A=row index (outside table), B–F=table
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 13

    # ── Row 1: From / Haulier ──
    _set(ws, "B1", "FROM BRITISH WOOL (BROMYARD)", bold=True)
    ws.merge_cells("B1:D1")
    _set(ws, "F1", f"Haulier: {haulier}", bold=True)

    # ── Row 2: Address ──
    _set(ws, "B2", "IIs Porthouse Industrial Estate Bromyard")
    ws.merge_cells("B2:F2")

    # ── Row 3: Destination ──
    _set(ws, "B3", "TO: BRITISH WOOL - BRECON DEPOT", bold=True)
    ws.merge_cells("B3:F3")

    # ── Rows 4–5: blank spacers ──

    # ── Row 6: DATE / Load N / PACKING SHEET ──
    _set(ws, "B6", "DATE",         bold=True,  border=True, halign="center")
    _set(ws, "C6", date_str,       fill=YELLOW, border=True, halign="center")
    _set(ws, "D6", f"Load {load_number}", fill=YELLOW, border=True, halign="center")
    ws.merge_cells("E6:F6")
    _set(ws, "E6", "PACKING SHEET", bold=True, size=14, border=True,
         halign="center", valign="center")

    # ── Row 7: blank ──

    # ── Row 8: Details heading ──
    ws.merge_cells("B8:F8")
    _set(ws, "B8", "DETAILS OF PACKS ON LOAD",
         bold=True, border=True, halign="center")

    # ── Row 9: blank ──

    # ── Row 10: Column headers ──
    for col, header in zip(
        ["B", "C", "D", "E", "F"],
        ["BALE NO", "Gross Weight", "COMMENTS", "NET weight", "No of wool"]
    ):
        _set(ws, f"{col}10", header,
             bold=True, border=True, halign="center", wrap=True)

    # ── Data rows ──
    DATA_START = 11
    for idx, row in enumerate(rows):
        r = DATA_START + idx

        # Row number outside table — no border
        ws[f"A{r}"].value = idx + 1
        ws[f"A{r}"].alignment = Alignment(horizontal="right")

        raw_no = row["bale_no"]
        ws[f"B{r}"].value = int(raw_no) if raw_no.isdigit() else raw_no
        ws[f"C{r}"].value = row["gross"]
        ws[f"D{r}"].value = row["comments"]
        ws[f"E{r}"].value = row["net"]
        ws[f"F{r}"].value = row["bags"]

        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{r}"].border = BORDER
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center")
        # Comments left-aligned for readability
        ws[f"D{r}"].alignment = Alignment(horizontal="left")

    # ── Totals row ──
    total_row = DATA_START + len(rows) + 1
    gross_total = sum(r["gross"] for r in rows)
    net_total   = sum(r["net"]   for r in rows)
    bags_total  = sum(r["bags"]  for r in rows)

    _set(ws, f"C{total_row}", gross_total, bold=True, border=True, halign="center")
    _set(ws, f"D{total_row}", "NET weight on load",
         bold=True, border=True, halign="center")
    _set(ws, f"E{total_row}", net_total,  bold=True, border=True, halign="center")
    _set(ws, f"F{total_row}", bags_total, bold=True, border=True, halign="center")

    # ── Summary box ──
    running_total = prev_total + net_total
    s = total_row + 2  # summary start row

    for offset, label, value in [
        (0, "WEIGHT ON LOAD", net_total),
        (1, "PREVIOUS TOT",   prev_total),
        (2, "RUNNING TOT",    running_total),
    ]:
        _set(ws, f"D{s+offset}", label,  fill=GREY,   bold=True, border=True, halign="center")
        _set(ws, f"E{s+offset}", value,  fill=YELLOW, border=True, halign="center")

    # Bales count — outside table, bottom-left of summary
    ws[f"A{s+2}"].value = "BALES"
    ws[f"A{s+2}"].font  = Font(bold=True)
    ws[f"A{s+2}"].alignment = Alignment(horizontal="right")
    ws[f"B{s+2}"].value = len(rows)
    ws[f"B{s+2}"].border = BORDER
    ws[f"B{s+2}"].alignment = Alignment(horizontal="center")

    # ── Footer ──
    footer_row = s + 4
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    _set(ws, f"A{footer_row}",
         "Category 3 animal by-products, not for human consumption",
         italic=True, halign="center")

    wb.save(output_path)

    # Summary to stdout for Claude to relay
    print(f"OK Packing sheet saved: {output_path}")
    print(f"   Load {load_number} | {len(rows)} bales | "
          f"Gross: {gross_total:.1f}kg | Net: {net_total:.1f}kg | Bags: {bags_total}")
    print(f"   Running total: {prev_total:.1f}kg + {net_total:.1f}kg = {running_total:.1f}kg")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Bromyard haulier packing sheet")
    parser.add_argument("--manifest",    required=True,              help="Depot suite CSV")
    parser.add_argument("--bale-log",    required=True,              help="bale-log.csv path")
    parser.add_argument("--load-number", required=True, type=int,    help="Load number")
    parser.add_argument("--haulier",     required=True,              help="Haulier name")
    parser.add_argument("--prev-total",  required=True, type=float,  help="Previous running total (kg)")
    parser.add_argument("--date",        required=True,              help="Date DD/MM/YYYY")
    parser.add_argument("--output",      required=True,              help="Output .xlsx path")
    args = parser.parse_args()

    bale_log      = load_bale_log(args.bale_log)
    manifest_rows = load_manifest(args.manifest)

    valid_rows, errors = cross_reference(manifest_rows, bale_log)

    hard_errors = [e for e in errors if e[1] == "HARD"]
    skip_errors = [e for e in errors if e[1] == "SKIP"]

    for bale_no, _, msg in skip_errors:
        print(f"  SKIPPED bale {bale_no}: {msg}", file=sys.stderr)

    if hard_errors:
        print("\nERRORS — resolve these before generating the packing sheet:", file=sys.stderr)
        for bale_no, _, msg in hard_errors:
            print(f"  ! Bale {bale_no}: {msg}", file=sys.stderr)
        sys.exit(1)

    if not valid_rows:
        print("No valid bales to process.", file=sys.stderr)
        sys.exit(1)

    valid_rows = build_comments(valid_rows)
    generate_xlsx(
        rows        = valid_rows,
        load_number = args.load_number,
        haulier     = args.haulier,
        prev_total  = args.prev_total,
        date_str    = args.date,
        output_path = args.output,
    )


if __name__ == "__main__":
    main()
